"""
Export functionality for timetables (PDF, Excel, CSV).
"""
from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import csv

from .models import Teacher, ClassGroup, Room, TimeSlot, TimetableEntry


def export_to_pdf(view_type, object_id=None):
    """
    Export timetable to PDF using ReportLab.
    """
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="timetable_{view_type}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4))
    elements = []
    styles = getSampleStyleSheet()

    # Title
    if view_type == 'master':
        title = "Master Timetable"
    elif view_type == 'teacher' and object_id:
        teacher = Teacher.objects.get(id=object_id)
        title = f"Timetable for {teacher.name}"
    elif view_type == 'class' and object_id:
        classgroup = ClassGroup.objects.get(id=object_id)
        title = f"Timetable for {classgroup.name}"
    elif view_type == 'room' and object_id:
        room = Room.objects.get(id=object_id)
        title = f"Timetable for {room.name}"
    else:
        title = "Timetable"

    elements.append(Paragraph(title, styles['Title']))
    elements.append(Spacer(1, 0.3 * inch))

    # Get data
    timeslots = TimeSlot.objects.all().order_by('day_index', 'period_index')

    # Group by day
    days = {}
    for ts in timeslots:
        if ts.day_index not in days:
            days[ts.day_index] = []
        days[ts.day_index].append(ts)

    # Build table
    for day_idx, day_slots in days.items():
        day_name = dict(TimeSlot.DAY_CHOICES)[day_idx]

        # Table header
        data = [[day_name] + [f"P{ts.period_index + 1}\n{ts.start_time.strftime('%H:%M')}" for ts in day_slots]]

        # Get entries for this day
        if view_type == 'master':
            classes = ClassGroup.objects.all()
            for classgroup in classes:
                row = [classgroup.name]
                for ts in day_slots:
                    entry = TimetableEntry.objects.filter(
                        classgroup=classgroup,
                        timeslot=ts
                    ).select_related('teacher', 'subject', 'room').first()

                    if entry:
                        cell_text = f"{entry.subject.name}\n{entry.teacher.name}\n{entry.room.name}"
                    else:
                        cell_text = "-"
                    row.append(cell_text)
                data.append(row)

        elif view_type == 'teacher' and object_id:
            teacher = Teacher.objects.get(id=object_id)
            row = ['Schedule']
            for ts in day_slots:
                entry = TimetableEntry.objects.filter(
                    teacher=teacher,
                    timeslot=ts
                ).select_related('classgroup', 'subject', 'room').first()

                if entry:
                    cell_text = f"{entry.subject.name}\n{entry.classgroup.name}\n{entry.room.name}"
                else:
                    cell_text = "-"
                row.append(cell_text)
            data.append(row)

        elif view_type == 'class' and object_id:
            classgroup = ClassGroup.objects.get(id=object_id)
            row = ['Schedule']
            for ts in day_slots:
                entry = TimetableEntry.objects.filter(
                    classgroup=classgroup,
                    timeslot=ts
                ).select_related('teacher', 'subject', 'room').first()

                if entry:
                    cell_text = f"{entry.subject.name}\n{entry.teacher.name}\n{entry.room.name}"
                else:
                    cell_text = "-"
                row.append(cell_text)
            data.append(row)

        elif view_type == 'room' and object_id:
            room = Room.objects.get(id=object_id)
            row = ['Schedule']
            for ts in day_slots:
                entry = TimetableEntry.objects.filter(
                    room=room,
                    timeslot=ts
                ).select_related('teacher', 'subject', 'classgroup').first()

                if entry:
                    cell_text = f"{entry.subject.name}\n{entry.teacher.name}\n{entry.classgroup.name}"
                else:
                    cell_text = "-"
                row.append(cell_text)
            data.append(row)

        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))

        elements.append(table)
        elements.append(Spacer(1, 0.5 * inch))

    doc.build(elements)
    return response


def export_to_excel(view_type, object_id=None):
    """
    Export timetable to Excel using openpyxl.
    """
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="timetable_{view_type}.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active

    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Title
    if view_type == 'master':
        ws.title = "Master Timetable"
    elif view_type == 'teacher' and object_id:
        teacher = Teacher.objects.get(id=object_id)
        ws.title = f"Teacher {object_id}"
    elif view_type == 'class' and object_id:
        classgroup = ClassGroup.objects.get(id=object_id)
        ws.title = f"Class {classgroup.name}"
    elif view_type == 'room' and object_id:
        room = Room.objects.get(id=object_id)
        ws.title = f"Room {room.name}"

    # Get data
    timeslots = TimeSlot.objects.all().order_by('day_index', 'period_index')
    days = {}
    for ts in timeslots:
        if ts.day_index not in days:
            days[ts.day_index] = []
        days[ts.day_index].append(ts)

    row_idx = 1

    for day_idx, day_slots in days.items():
        day_name = dict(TimeSlot.DAY_CHOICES)[day_idx]

        # Day header
        ws.cell(row=row_idx, column=1, value=day_name)
        ws.cell(row=row_idx, column=1).font = header_font
        ws.cell(row=row_idx, column=1).fill = header_fill
        ws.cell(row=row_idx, column=1).alignment = center_align

        # Period headers
        for col_idx, ts in enumerate(day_slots, start=2):
            cell_value = f"P{ts.period_index + 1}\n{ts.start_time.strftime('%H:%M')}"
            ws.cell(row=row_idx, column=col_idx, value=cell_value)
            ws.cell(row=row_idx, column=col_idx).font = header_font
            ws.cell(row=row_idx, column=col_idx).fill = header_fill
            ws.cell(row=row_idx, column=col_idx).alignment = center_align

        row_idx += 1

        # Data rows
        if view_type == 'master':
            classes = ClassGroup.objects.all()
            for classgroup in classes:
                ws.cell(row=row_idx, column=1, value=classgroup.name)
                for col_idx, ts in enumerate(day_slots, start=2):
                    entry = TimetableEntry.objects.filter(
                        classgroup=classgroup,
                        timeslot=ts
                    ).select_related('teacher', 'subject', 'room').first()

                    if entry:
                        cell_value = f"{entry.subject.name}\n{entry.teacher.name}\n{entry.room.name}"
                    else:
                        cell_value = "-"
                    ws.cell(row=row_idx, column=col_idx, value=cell_value)
                    ws.cell(row=row_idx, column=col_idx).alignment = center_align
                row_idx += 1

        elif view_type == 'teacher' and object_id:
            teacher = Teacher.objects.get(id=object_id)
            ws.cell(row=row_idx, column=1, value='Schedule')
            for col_idx, ts in enumerate(day_slots, start=2):
                entry = TimetableEntry.objects.filter(
                    teacher=teacher,
                    timeslot=ts
                ).select_related('classgroup', 'subject', 'room').first()

                if entry:
                    cell_value = f"{entry.subject.name}\n{entry.classgroup.name}\n{entry.room.name}"
                else:
                    cell_value = "-"
                ws.cell(row=row_idx, column=col_idx, value=cell_value)
                ws.cell(row=row_idx, column=col_idx).alignment = center_align
            row_idx += 1

        elif view_type == 'class' and object_id:
            classgroup = ClassGroup.objects.get(id=object_id)
            ws.cell(row=row_idx, column=1, value='Schedule')
            for col_idx, ts in enumerate(day_slots, start=2):
                entry = TimetableEntry.objects.filter(
                    classgroup=classgroup,
                    timeslot=ts
                ).select_related('teacher', 'subject', 'room').first()

                if entry:
                    cell_value = f"{entry.subject.name}\n{entry.teacher.name}\n{entry.room.name}"
                else:
                    cell_value = "-"
                ws.cell(row=row_idx, column=col_idx, value=cell_value)
                ws.cell(row=row_idx, column=col_idx).alignment = center_align
            row_idx += 1

        elif view_type == 'room' and object_id:
            room = Room.objects.get(id=object_id)
            ws.cell(row=row_idx, column=1, value='Schedule')
            for col_idx, ts in enumerate(day_slots, start=2):
                entry = TimetableEntry.objects.filter(
                    room=room,
                    timeslot=ts
                ).select_related('teacher', 'subject', 'classgroup').first()

                if entry:
                    cell_value = f"{entry.subject.name}\n{entry.teacher.name}\n{entry.classgroup.name}"
                else:
                    cell_value = "-"
                ws.cell(row=row_idx, column=col_idx, value=cell_value)
                ws.cell(row=row_idx, column=col_idx).alignment = center_align
            row_idx += 1

        row_idx += 1  # Empty row between days

    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response


def export_to_csv(view_type, object_id=None):
    """
    Export timetable to CSV.
    """
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="timetable_{view_type}.csv"'

    writer = csv.writer(response)

    # Header
    writer.writerow(['Day', 'Period', 'Time', 'Class', 'Subject', 'Teacher', 'Room'])

    # Get entries
    if view_type == 'master':
        entries = TimetableEntry.objects.all()
    elif view_type == 'teacher' and object_id:
        entries = TimetableEntry.objects.filter(teacher_id=object_id)
    elif view_type == 'class' and object_id:
        entries = TimetableEntry.objects.filter(classgroup_id=object_id)
    elif view_type == 'room' and object_id:
        entries = TimetableEntry.objects.filter(room_id=object_id)
    else:
        entries = TimetableEntry.objects.all()

    entries = entries.select_related(
        'classgroup', 'subject', 'teacher', 'room', 'timeslot'
    ).order_by('timeslot__day_index', 'timeslot__period_index')

    for entry in entries:
        writer.writerow([
            entry.timeslot.get_day_index_display(),
            f"Period {entry.timeslot.period_index + 1}",
            f"{entry.timeslot.start_time} - {entry.timeslot.end_time}",
            entry.classgroup.name,
            entry.subject.name,
            entry.teacher.name,
            entry.room.name,
        ])

    return response
